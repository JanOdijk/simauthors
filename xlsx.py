from openpyxl import load_workbook
import xlsxwriter

eps = ''
defaultcolwidth = 15


def getxlsxdata(fullname, headerrow=0, sheetname=None):
    data = []
    header = []
    try:
        wb = load_workbook(fullname, data_only=True, read_only=True)
    except Exception:
        return (header, data)
    wsns = wb.sheetnames
    if wsns != []:
        if sheetname is None:
            ws = wb[wsns[0]]
        else:
            ws = wb[sheetname]
    else:
        print(f"Error reading {fullname}. No worksheets found")
    rowctr = -1
    for row in ws.iter_rows():
        rowctr += 1
        #preheader
        if rowctr < headerrow:
            continue
        # header
        elif rowctr == headerrow:
            header = [eps if cell.value is None else cell.value for cell in row]
        else:
            valuerow = [eps if cell.value is None else cell.value for cell in row]
            data.append(valuerow)
    wb.close()
    return header, data

def mkworkbook(outfullname, headers, allrows, sheetname='Sheet1', freeze_panes=None, formats=[]):
    workbook = xlsxwriter.Workbook(outfullname, {"strings_to_numbers": True})
    bold = workbook.add_format({'bold': True})

    realformats = []
    for fmt in formats:
        if fmt is None:
            realformats.append(fmt)
        else:
            realfmt = workbook.add_format(fmt)
            realformats.append(realfmt)

    worksheet1 = workbook.add_worksheet(sheetname)

    # worksheet1
    if freeze_panes is not None:
        (r, c) = freeze_panes
        worksheet1.freeze_panes(r, c)

    colctr = 0
    if headers != []:
        for val in headers[-1]:
            if val is None:
                cval = ''
            else:
                cval = str(val)
            colwidth = len(cval) if len(cval) > defaultcolwidth else defaultcolwidth
            worksheet1.set_column(colctr, colctr, colwidth)
            colctr += 1

    rowctr = 0
    for header in headers:
        xlsx_writerow(worksheet1, rowctr, header, format=bold)
        rowctr += 1

    for row in allrows:
        xlsx_writerow(worksheet1, rowctr, row, formats=realformats)
        rowctr += 1

    worksheet1.autofilter(0, 0, rowctr, colctr)
    return workbook


def xlsx_writerow(sheet, rowctr, row, format=None, formats=[]):
    lrow = len(row)
    lformats = len(formats)
    for colctr in range(lrow):
        if format is None:
            if formats != [] and colctr < lformats and formats[colctr] is not None:
                sheet.write(rowctr, colctr, row[colctr], formats[colctr])
            else:
                sheet.write(rowctr, colctr, row[colctr])
        else:
            sheet.write(rowctr, colctr, row[colctr], format)
